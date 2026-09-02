from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from kerui_recruit.org.fields import FLAT_COLUMNS, Column, client_columns


def _workbook(rows: list[dict[str, object]], columns: list[Column]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "组织架构"
    sheet.append([column.label for column in columns])

    for row in rows:
        sheet.append([row.get(column.key, "") for column in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_internal(rows: list[dict[str, object]]) -> bytes:
    """Full-fidelity export including sensitive fields."""
    return _workbook(rows, list(FLAT_COLUMNS))


def export_client(rows: list[dict[str, object]]) -> bytes:
    """Client-facing export with sensitive fields stripped and styled headers."""
    columns = client_columns()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "组织架构"

    header = [column.label for column in columns]
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        sheet.append([row.get(column.key, "") for column in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# 部门色系：不同顶层部门使用不同色相，同一部门及其子部门共用同一色系（淡色）。
_DEPT_PALETTE = [
    (0.88, 0.95, 0.90),  # 淡绿
    (0.99, 0.92, 0.85),  # 淡橙
    (0.93, 0.90, 0.97),  # 淡紫
    (0.99, 0.97, 0.84),  # 淡黄
    (0.85, 0.94, 0.93),  # 淡青
    (0.97, 0.90, 0.93),  # 淡粉
    (0.94, 0.97, 0.88),  # 淡黄绿
    (0.96, 0.92, 0.88),  # 淡棕
]
# 公司根节点使用淡蓝色，与部门色系区分。
_COMPANY_FILL = (0.85, 0.92, 0.99)
_COMPANY_STROKE = (0.50, 0.68, 0.88)
_COMPANY_TEXT = (0.08, 0.20, 0.40)


def _fangsong_font_path() -> str | None:
    """Locate a FangSong (仿宋) TTF on the host; return None to fall back to a built-in CJK font."""
    import os

    for path in (
        r"C:\Windows\Fonts\simfang.ttf",
        r"C:\Windows\Fonts\STFANGSO.TTF",
        "/System/Library/Fonts/STFangsong.ttf",
        "/Library/Fonts/FangSong.ttf",
        "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
    ):
        if os.path.exists(path):
            return path
    return None


def export_arch_pdf(root, *, orientation: str = "vertical", watermark: str = "") -> bytes:
    """Render a department-only org tree as a single-page architecture chart.

    - Department nodes show ``名称-负责人`` (``名称-XXX`` when no leader).
    - FangSong font, centered labels, and font/box sizes that adapt to the
      number of departments.
    - The company root uses a light-blue box; each top-level department (and
      its sub-departments) uses one light colour family, distinct per department.
    """
    import pymupdf
    from datetime import datetime

    # 扫描部门数量与最长标签，用于自适应字号/框大小。
    dept_count = 0
    max_chars = 4

    def scan(node) -> None:
        nonlocal dept_count, max_chars
        if node.kind == "department":
            dept_count += 1
            max_chars = max(max_chars, len(f"{node.name}-{node.leader_name or 'XXX'}"))
        for child in node.children:
            scan(child)

    scan(root)

    if dept_count <= 5:
        base_font = 13.0
    elif dept_count <= 12:
        base_font = 12.0
    elif dept_count <= 25:
        base_font = 11.0
    elif dept_count <= 50:
        base_font = 10.0
    else:
        base_font = 9.0

    node_w = max(110.0, base_font * max_chars * 1.15 + 16.0)
    node_h = base_font * 4.4
    gap = base_font * 2.2
    level = base_font * 5.0
    margin = 56.0

    def subtree_span(node) -> float:
        if not node.children:
            return node_w
        return max(node_w, sum(subtree_span(c) for c in node.children) + gap * (len(node.children) - 1))

    positions: dict[str, tuple[float, int]] = {}
    max_depth = 0

    def assign(node, depth: int, left: float) -> None:
        nonlocal max_depth
        span = subtree_span(node)
        positions[node.id] = (left + span / 2, depth)
        max_depth = max(max_depth, depth)
        child_left = left
        for child in node.children:
            child_span = subtree_span(child)
            assign(child, depth + 1, child_left)
            child_left += child_span + gap

    assign(root, 0, 0)

    horizontal = orientation == "horizontal"
    span = subtree_span(root)
    if horizontal:
        canvas_w = max_depth * (node_w + level) + node_w
        canvas_h = span
    else:
        canvas_w = span
        canvas_h = max_depth * (node_h + level) + node_h

    page_w, page_h = 842.0, 595.0
    usable_w = page_w - 2 * margin
    usable_h = page_h - 2 * margin
    scale = min(1.0, usable_w / canvas_w, usable_h / canvas_h)
    fontsize = max(6.0, base_font * scale)

    draw_w = canvas_w * scale
    draw_h = canvas_h * scale
    origin_x = margin + (usable_w - draw_w) / 2
    origin_y = margin + (usable_h - draw_h) / 2

    def coord(node):
        center, depth = positions[node.id]
        if horizontal:
            x = depth * (node_w + level) + node_w / 2
            y = center
        else:
            x = center
            y = depth * (node_h + level) + node_h / 2
        return (origin_x + x * scale, origin_y + y * scale)

    def node_rect(node):
        x, y = coord(node)
        w = node_w * scale
        h = node_h * scale
        return pymupdf.Rect(x - w / 2, y - h / 2, x + w / 2, y + h / 2)

    # 为每个顶层部门分配一个色系，子部门继承同一色系。
    hue_by_dept: dict[str, int] = {}
    hue_counter = [0]

    def build_hues(node, hue: int | None) -> None:
        if node.kind == "department":
            if hue is None:
                hue = hue_counter[0] % len(_DEPT_PALETTE)
                hue_counter[0] += 1
            hue_by_dept[node.id] = hue
        for child in node.children:
            build_hues(child, hue if node.kind == "department" else None)

    build_hues(root, None)

    def node_colors(node):
        if node.kind == "company":
            return _COMPANY_FILL, _COMPANY_STROKE, _COMPANY_TEXT
        fill = _DEPT_PALETTE[hue_by_dept[node.id]]
        stroke = tuple(min(1.0, c * 0.55) for c in fill)
        text = tuple(max(0.0, c * 0.28) for c in fill)
        return fill, stroke, text

    document = pymupdf.open()
    page = document.new_page(width=page_w, height=page_h)
    page.draw_rect(page.rect, color=None, fill=(0.985, 0.99, 0.987), width=0)

    # 加载仿宋字体（不可用则回退到内置 CJK 字体）。
    font_name = "china-s"
    font = None
    fang_path = _fangsong_font_path()
    if fang_path:
        try:
            page.insert_font(fontname="fangsong", fontfile=fang_path)
            font = pymupdf.Font(fontfile=fang_path)
            font_name = "fangsong"
        except Exception:
            font = None
            font_name = "china-s"

    def text_width(text: str, size: float) -> float:
        if font is not None:
            return font.text_length(text, size)
        return pymupdf.get_text_length(text, fontname="china-s", fontsize=size)

    edge = (0.72, 0.76, 0.73)

    page.insert_text((margin, margin - 24), root.name, fontsize=max(12.0, base_font * 1.3), fontname=font_name)
    if watermark:
        page.insert_textbox(
            page.rect,
            watermark,
            fontsize=40,
            fontname=font_name,
            align=pymupdf.TEXT_ALIGN_CENTER,
            color=(0.88, 0.91, 0.89),
        )
    page.insert_text(
        (margin, page_h - 16),
        f"生成日期：{datetime.now().strftime('%Y-%m-%d')}",
        fontsize=8,
        fontname=font_name,
    )

    def draw(node) -> None:
        rect = node_rect(node)
        for child in node.children:
            crect = node_rect(child)
            if horizontal:
                mid_x = (rect.x1 + crect.x0) / 2
                page.draw_line((rect.x1, rect.y0 + rect.height / 2), (mid_x, rect.y0 + rect.height / 2), color=edge, width=1)
                page.draw_line((mid_x, rect.y0 + rect.height / 2), (mid_x, crect.y0 + crect.height / 2), color=edge, width=1)
                page.draw_line((mid_x, crect.y0 + crect.height / 2), (crect.x0, crect.y0 + crect.height / 2), color=edge, width=1)
            else:
                mid_y = (rect.y1 + crect.y0) / 2
                page.draw_line((rect.x0 + rect.width / 2, rect.y1), (rect.x0 + rect.width / 2, mid_y), color=edge, width=1)
                page.draw_line((rect.x0 + rect.width / 2, mid_y), (crect.x0 + crect.width / 2, mid_y), color=edge, width=1)
                page.draw_line((crect.x0 + crect.width / 2, mid_y), (crect.x0 + crect.width / 2, crect.y0), color=edge, width=1)
            draw(child)

        fill, stroke, text_color = node_colors(node)
        page.draw_rect(rect, color=stroke, fill=fill, width=1)

        label = node.name
        if node.kind == "department":
            label = f"{node.name}-{node.leader_name or 'XXX'}"

        # 水平 + 垂直居中。
        tx = (rect.x0 + rect.x1) / 2 - text_width(label, fontsize) / 2
        ty = (rect.y0 + rect.y1) / 2 + fontsize * 0.35
        page.insert_text((tx, ty), label, fontsize=fontsize, fontname=font_name, color=text_color)

    draw(root)

    payload = document.tobytes()
    document.close()
    return payload
